import openpyxl


# class State:
#     def __init__(self) -> None:
#         self.complete = False
#         self.timeout = 10

# class Init(State):
#     def __init__(self) -> None:
#         super().__init__()
#         tu = TestUnit()

class TestUnit(object):
    def __init__(self) -> None:
        self.sheet = self.createExcelSheet(path='./ascii-table.xlsx')               # Enter Path here
        self.list = self.fillList(num_column=2, char_column=3, sheet=self.sheet)    # Enter the columns here
        self.passed = None
        self.inNum = -1
        self.outChar = ''

    def createExcelSheet(self, path):
        workbook = openpyxl.load_workbook(filename=path, data_only=True)            
        sheet = workbook.active
        print(workbook.sheetnames[0])
        print(sheet.title)
        return sheet

    def fillList(self,num_column, char_column, sheet):
        self.row_cnt = sheet.max_row
        print(f"Max Rows: {self.row_cnt}")
        liste = {}

        for row in range(1, self.row_cnt):
            c = sheet.cell(row=row, column=char_column)
            n = sheet.cell(row=row, column=num_column)
            c = c.value
            n = n.value
            if n is not None and n.isdigit():
                liste[c] = int(n)

        return liste


    def runTest(self) -> None:
        pass


tu = TestUnit()
tu.runTest()
print(tu.list)