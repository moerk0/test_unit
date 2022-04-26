from mimetypes import guess_type

from openpyxl import Workbook, cell, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color

from test_data import OutLevel, ReturnCode, TestData, log


class Excel:
    def __init__(
        self,
        lang: str,
        filename="./data/sprachtabelle.xlsx",
        num=0,
        brai=1,
        cha=2,
        key=3,
        mod=4,
        dead=5,
        dmod=6,
        rec=8,
        pas=9,
    ) -> None:

        # In Section
        self.filename = filename
        self.workbook_origin = load_workbook(filename=filename, data_only=True)

        self.columns = {
            "num": num,
            "brai": brai,
            "cha": cha,
            "key": key,
            "mod": mod,
            "dead": dead,
            "dmod": dmod,
            "rec": rec,
            # 'passed': pas
        }

        # Language Section
        self.languages = {  # TODO: Automate recognition process
            "US-ENGLISCH": 0,
            "FRANZÖSISCH": 1,
            "SPANISCH": 2,
            "ITALIENISCH": 3,
            "TÜRKISCH": 4,
            "DEUTSCH": 5,
        }

        self.lang = lang.upper()
        self.workbook_origin.active = self.languages[self.lang]  # type: ignore

        # Out Section
        self.workbook_result = None  # createresultfile()
        # self.outColumns = self.setColumns()             # Change rec and pass col here
        # log.debug(f"OutColumn values are: {self.outColumns}")             # In which cols will the results be?
        self.createBackup()
        self.outFile = self.filename  #               f'{filename[:-5]}_results.xlsx'

    def setColumns(self):
        """Use this function to order the Columns of the out file from 1 to n.
        This is nessecary because you may change the order of the entries,
        and in some languages there are more columns than in others."""

        d = {}
        for i, keys in enumerate(self.columns):
            d[keys] = i
        return d

    def verifyColumns(*cols):
        if "Alias" in cols or None in cols:
            log.error(" Please specify the correct columns")
            exit()

    def getTestData(self, single_row=None) -> "list[TestData]":
        """If you cast single row an integer value you will only get the specified row"""

        row_cnt = self.workbook_origin.active.max_row
        log.info(f"Max Rows: {row_cnt}")
        l = []

        if single_row == None:
            rows = self.workbook_origin.active.iter_rows(min_row=1, max_row=row_cnt)
        else:
            rows = self.workbook_origin.active.iter_rows(
                min_row=single_row, max_row=single_row
            )

        for row in rows:

            n = row[self.columns["num"]].value
            b = row[self.columns["brai"]].value
            c = row[self.columns["cha"]].value

            k = row[self.columns["key"]].value
            m = row[self.columns["mod"]].value

            d = row[self.columns["dead"]].value

            dm = row[self.columns["dmod"]].value

            if row[0].row == 1:  # This displays the names of the columns
                print(
                    f"Column names: Bin->{n}  Brai->{b} cha->{c}  key->{k}  mod->{m}  deadkey->{d}  dmod->{dm}"
                )
                self.verifyColumns(n, c, b, k, m, d, dm)

            else:

                try:
                    d = TestData(row[0].row, int(n), b, c, k, m, d, dm)
                    if n != 0:
                        l.append(d)
                    else:
                        log.warning(f"row: {row[0].row} - {n} is NULL, skipped")
                except:
                    log.warning(f"row: {row[0].row} - {n} is no integer, skipped")

                log.debug(f"{d.content(OutLevel.INITIAL)}")

        return l

    def drawColors(self, *cels: list):
        colors = {
            ReturnCode.BLANK.value: PatternFill(fill_type=None),
            ReturnCode.GREEN.value: PatternFill(
                start_color="00FF00", patternType="solid"
            ),
            ReturnCode.YELLOW.value: PatternFill(
                start_color="FFFF00", patternType="solid"
            ),
            ReturnCode.RED.value: PatternFill(
                start_color="FF0000", patternType="solid"
            ),
        }

        try:
            for cel in cels:
                cel[0].fill = colors[cel[1]]

        except:
            log.error(
                f"{cel} is {type(cel)}. Required Format: list [cell.obj, int(color_code)]"
            )

    def writeData(self, d: "list[TestData]"):
        len_res = len(d)

        for i in range(len_res):
            data: TestData = d[i]

            cn = self.workbook_origin.active.cell(
                row=d[i].row, column=self.columns["num"] + 1
            )
            cc = self.workbook_origin.active.cell(
                row=data.row, column=self.columns["cha"] + 1
            )
            ck = self.workbook_origin.active.cell(
                row=data.row, column=self.columns["key"] + 1
            )
            cm = self.workbook_origin.active.cell(
                row=data.row, column=self.columns["mod"] + 1
            )
            cd = self.workbook_origin.active.cell(
                row=data.row, column=self.columns["dead"] + 1
            )
            cdm = self.workbook_origin.active.cell(
                row=data.row, column=self.columns["dmod"] + 1
            )

            self.drawColors(
                [cn, data.color_num.value],
                [cc, data.color_cha.value],
                [ck, data.color_key.value],
                [cm, data.color_mod.value],
                [cd, data.color_ded.value],
                [cdm, data.color_dmo.value],
            )

            cc.value = data.cha
            ck.value = data.key
            cm.value = data.mod
            cd.value = data.dead
            cdm.value = data.dmod

            # row[self.outColumns['dmod']].value = data.dmod

    def writeResults(self, d: list):
        """
        Always call createResultFile before writing the results,
        change result columns here
        """
        len_res = len(d)

        for i in range(len_res):
            data: TestData = d[i]

            cr = self.workbook_origin.active.cell(
                row=data.row, column=self.columns["rec"] + 1
            )
            # cp = self.workbook_origin.active.cell(row=data.row, column=self.columns['passed']+1)
            cr.value = data.received
            # cp.value = data.passed

            # color for passed
            if data.passed is True:
                data.color_rec = ReturnCode.GREEN
            elif data.received is None:
                data.color_rec = ReturnCode.YELLOW
            else:
                data.color_rec = ReturnCode.RED
            self.drawColors([cr, data.color_rec.value])

    def createBackup(self):
        bu_name = f"{self.filename[:-5]}_bu.xlsx"
        try:
            load_workbook(bu_name)
            log.info("BackUp File Found")
        except:
            self.workbook_origin.save(filename=bu_name)
            log.info("backup file created")

    def saveResultFile(self):
        # self.workbook_result.save(filename= self.outFile)
        self.workbook_origin.save(filename=self.filename)
        log.info("File saved")


# ex = Excel("französisch", "./data/sprachtabelle2.xlsx")
# d = ex.getTestData(71)
# # # print(len(d))
# for data in d:
#      data.validate_all()


# # # ex.createResultFile()
# ex.writeData(d)
# ex.writeResults(d)
# ex.saveResultFile()
